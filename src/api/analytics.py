from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func, case, and_, or_, desc
from typing import List, Optional
from pydantic import BaseModel, Field, ConfigDict
from datetime import datetime, date
from decimal import Decimal
import math
import io
import csv

from src.db import get_db
from src.db.models import (
    Movement, Item, Lot, Material, PurchaseOrder, PurchaseOrderItem,
    MaterialShape, MovementType, MaterialGroup, MaterialGroupMember
)

router = APIRouter()

# ========================================
# Pydanticスキーマ
# ========================================

class AnalyticsSummaryRequest(BaseModel):
    """集計検索条件"""
    start_date: Optional[date] = Field(None, description="開始日")
    end_date: Optional[date] = Field(None, description="終了日")
    material_name: Optional[str] = Field(None, description="材料名（部分一致）")
    material_group_id: Optional[int] = Field(None, description="材料グループID")
    purchase_month: Optional[str] = Field(None, description="購入月（YYMM形式）")
    supplier: Optional[str] = Field(None, description="仕入先（部分一致）")
    movement_type: Optional[MovementType] = Field(None, description="入出庫種別")

class MaterialSummary(BaseModel):
    """材料別集計データ"""
    model_config = ConfigDict(from_attributes=True)

    material_id: int
    material_name: str
    current_stock_quantity: int = 0  # 現在在庫本数
    current_stock_weight_kg: float = 0.0  # 現在在庫重量
    total_in_quantity: int = 0  # 入庫本数
    total_in_weight_kg: float = 0.0  # 入庫重量
    total_out_quantity: int = 0  # 出庫本数
    total_out_weight_kg: float = 0.0  # 出庫重量
    total_amount: float = 0.0  # 合計金額（入庫時単価×数量）

class AnalyticsSummaryResponse(BaseModel):
    """集計結果レスポンス"""
    materials: List[MaterialSummary]
    total_stock_quantity: int
    total_stock_weight_kg: float
    total_in_quantity: int
    total_in_weight_kg: float
    total_out_quantity: int
    total_out_weight_kg: float
    total_amount: float

class GraphDataPoint(BaseModel):
    """グラフデータポイント"""
    label: str
    value: float

class GraphDataResponse(BaseModel):
    """グラフデータレスポンス"""
    labels: List[str]
    datasets: List[dict]

# ========================================
# ユーティリティ関数
# ========================================

def _calculate_weight_kg(material: Material, length_mm: int, quantity: int) -> float:
    """重量計算（kg）"""
    if not material or quantity <= 0:
        return 0.0

    length_cm = length_mm / 10

    if material.shape == MaterialShape.ROUND:
        radius_cm = (material.diameter_mm / 2) / 10
        volume_cm3 = math.pi * (radius_cm ** 2) * length_cm
    elif material.shape == MaterialShape.HEXAGON:
        side_cm = (material.diameter_mm / 2) / 10
        volume_cm3 = (3 * math.sqrt(3) / 2) * (side_cm ** 2) * length_cm
    elif material.shape == MaterialShape.SQUARE:
        side_cm = material.diameter_mm / 10
        volume_cm3 = (side_cm ** 2) * length_cm
    else:
        volume_cm3 = 0.0

    weight_g = volume_cm3 * material.current_density
    return round((weight_g * quantity) / 1000, 3)

# ========================================
# APIエンドポイント
# ========================================

@router.get("/summary/", response_model=AnalyticsSummaryResponse)
async def get_analytics_summary(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    material_name: Optional[str] = Query(None),
    material_group_id: Optional[int] = Query(None),
    purchase_month: Optional[str] = Query(None),
    supplier: Optional[str] = Query(None),
    movement_type: Optional[MovementType] = Query(None),
    db: Session = Depends(get_db)
):
    """集計検索API

    検索条件に基づいて材料別の在庫数、入出庫数、金額を集計します。
    """
    # 材料フィルタ構築
    material_query = db.query(Material.id).filter(Material.is_active == True)

    if material_name:
        material_query = material_query.filter(Material.display_name.contains(material_name))

    if material_group_id:
        material_query = material_query.join(MaterialGroupMember).filter(
            MaterialGroupMember.group_id == material_group_id
        )

    material_ids = [m[0] for m in material_query.all()]

    if not material_ids:
        return AnalyticsSummaryResponse(
            materials=[],
            total_stock_quantity=0,
            total_stock_weight_kg=0.0,
            total_in_quantity=0,
            total_in_weight_kg=0.0,
            total_out_quantity=0,
            total_out_weight_kg=0.0,
            total_amount=0.0
        )

    # ロット・アイテムのフィルタ条件
    lot_filters = [Lot.material_id.in_(material_ids)]

    if purchase_month:
        lot_filters.append(Lot.purchase_month == purchase_month)

    if supplier:
        lot_filters.append(Lot.supplier.contains(supplier))

    if start_date:
        lot_filters.append(Lot.received_date >= datetime.combine(start_date, datetime.min.time()))

    if end_date:
        lot_filters.append(Lot.received_date <= datetime.combine(end_date, datetime.max.time()))

    # 材料別に集計
    materials_summary = []
    total_stock_qty = 0
    total_stock_weight = 0.0
    total_in_qty = 0
    total_in_weight = 0.0
    total_out_qty = 0
    total_out_weight = 0.0
    total_amount_sum = 0.0

    for material_id in material_ids:
        material = db.query(Material).filter(Material.id == material_id).first()
        if not material:
            continue

        # 現在在庫（条件に一致するロットのアイテム）
        current_items = db.query(Item).join(Lot).filter(
            and_(
                *lot_filters,
                Lot.material_id == material_id,
                Item.is_active == True
            )
        ).all()

        current_qty = sum(item.current_quantity for item in current_items)
        current_weight = sum(
            _calculate_weight_kg(material, item.lot.length_mm, item.current_quantity)
            for item in current_items
        )

        # 入出庫履歴（条件に一致するロットのアイテムの履歴）
        item_ids = [item.id for item in current_items]
        movement_filters = [Movement.item_id.in_(item_ids)] if item_ids else []

        if movement_type:
            movement_filters.append(Movement.movement_type == movement_type)

        if start_date:
            movement_filters.append(Movement.processed_at >= datetime.combine(start_date, datetime.min.time()))

        if end_date:
            movement_filters.append(Movement.processed_at <= datetime.combine(end_date, datetime.max.time()))

        if movement_filters:
            movements = db.query(Movement).filter(and_(*movement_filters)).all()
        else:
            movements = []

        in_qty = sum(m.quantity for m in movements if m.movement_type == MovementType.IN)
        out_qty = sum(m.quantity for m in movements if m.movement_type == MovementType.OUT)

        # 重量計算（各移動の所属アイテムから長さを取得）
        in_weight = 0.0
        out_weight = 0.0
        for m in movements:
            item = db.query(Item).filter(Item.id == m.item_id).first()
            if item and item.lot:
                weight = _calculate_weight_kg(material, item.lot.length_mm, m.quantity)
                if m.movement_type == MovementType.IN:
                    in_weight += weight
                else:
                    out_weight += weight

        # 金額計算（ロットの入庫時金額を集計）
        lots = db.query(Lot).filter(and_(*lot_filters, Lot.material_id == material_id)).all()
        amount = sum(lot.received_amount or 0.0 for lot in lots)

        materials_summary.append(MaterialSummary(
            material_id=material.id,
            material_name=material.display_name,
            current_stock_quantity=current_qty,
            current_stock_weight_kg=round(current_weight, 3),
            total_in_quantity=in_qty,
            total_in_weight_kg=round(in_weight, 3),
            total_out_quantity=out_qty,
            total_out_weight_kg=round(out_weight, 3),
            total_amount=round(amount, 2)
        ))

        total_stock_qty += current_qty
        total_stock_weight += current_weight
        total_in_qty += in_qty
        total_in_weight += in_weight
        total_out_qty += out_qty
        total_out_weight += out_weight
        total_amount_sum += amount

    return AnalyticsSummaryResponse(
        materials=materials_summary,
        total_stock_quantity=total_stock_qty,
        total_stock_weight_kg=round(total_stock_weight, 3),
        total_in_quantity=total_in_qty,
        total_in_weight_kg=round(total_in_weight, 3),
        total_out_quantity=total_out_qty,
        total_out_weight_kg=round(total_out_weight, 3),
        total_amount=round(total_amount_sum, 2)
    )

@router.get("/graph/timeline/", response_model=GraphDataResponse)
async def get_timeline_graph(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    material_id: Optional[int] = Query(None),
    db: Session = Depends(get_db)
):
    """時系列推移グラフデータ（日別の入出庫推移）"""
    query = db.query(
        func.date(Movement.processed_at).label('date'),
        Movement.movement_type,
        func.sum(Movement.quantity).label('total_quantity')
    ).join(Item).join(Lot)

    if start_date:
        query = query.filter(Movement.processed_at >= datetime.combine(start_date, datetime.min.time()))

    if end_date:
        query = query.filter(Movement.processed_at <= datetime.combine(end_date, datetime.max.time()))

    if material_id:
        query = query.filter(Lot.material_id == material_id)

    results = query.group_by(func.date(Movement.processed_at), Movement.movement_type).order_by('date').all()

    # データ整形
    dates = sorted(set(r.date for r in results))
    labels = [d.strftime('%Y-%m-%d') for d in dates]

    in_data = []
    out_data = []

    for d in dates:
        in_qty = sum(r.total_quantity for r in results if r.date == d and r.movement_type == MovementType.IN)
        out_qty = sum(r.total_quantity for r in results if r.date == d and r.movement_type == MovementType.OUT)
        in_data.append(in_qty)
        out_data.append(out_qty)

    return GraphDataResponse(
        labels=labels,
        datasets=[
            {"label": "入庫", "data": in_data, "borderColor": "rgb(34, 197, 94)", "backgroundColor": "rgba(34, 197, 94, 0.2)"},
            {"label": "出庫", "data": out_data, "borderColor": "rgb(239, 68, 68)", "backgroundColor": "rgba(239, 68, 68, 0.2)"}
        ]
    )

@router.get("/graph/material-composition/", response_model=GraphDataResponse)
async def get_material_composition_graph(
    db: Session = Depends(get_db)
):
    """材料別構成比グラフデータ（円グラフ）"""
    try:
        query = db.query(
            Material.display_name,
            func.sum(Item.current_quantity).label('total_quantity')
        ).join(Lot, Material.id == Lot.material_id).join(Item, Lot.id == Item.lot_id).filter(
            Material.is_active == True,
            Item.is_active == True
        ).group_by(Material.id, Material.display_name).order_by(desc('total_quantity')).limit(10)

        results = query.all()

        labels = [r.display_name for r in results]
        data = [int(r.total_quantity) for r in results]

        return GraphDataResponse(
            labels=labels,
            datasets=[{
                "label": "在庫本数",
                "data": data,
                "backgroundColor": [
                    'rgb(59, 130, 246)', 'rgb(34, 197, 94)', 'rgb(251, 146, 60)',
                    'rgb(168, 85, 247)', 'rgb(236, 72, 153)', 'rgb(14, 165, 233)',
                    'rgb(132, 204, 22)', 'rgb(251, 191, 36)', 'rgb(244, 63, 94)',
                    'rgb(99, 102, 241)'
                ]
            }]
        )
    except Exception as e:
        import traceback
        print(f"材料別構成比グラフエラー: {traceback.format_exc()}")
        # エラー時は空データを返す
        return GraphDataResponse(labels=[], datasets=[{"label": "在庫本数", "data": [], "backgroundColor": []}])

@router.get("/graph/supplier-amount/", response_model=GraphDataResponse)
async def get_supplier_amount_graph(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: Session = Depends(get_db)
):
    """仕入先別金額グラフデータ（棒グラフ）"""
    query = db.query(
        Lot.supplier,
        func.sum(Lot.received_amount).label('total_amount')
    ).filter(Lot.supplier.isnot(None))

    if start_date:
        query = query.filter(Lot.received_date >= datetime.combine(start_date, datetime.min.time()))

    if end_date:
        query = query.filter(Lot.received_date <= datetime.combine(end_date, datetime.max.time()))

    results = query.group_by(Lot.supplier).order_by(desc('total_amount')).limit(10).all()

    labels = [r.supplier for r in results]
    data = [float(r.total_amount or 0) for r in results]

    return GraphDataResponse(
        labels=labels,
        datasets=[{
            "label": "金額（円）",
            "data": data,
            "backgroundColor": "rgba(59, 130, 246, 0.8)"
        }]
    )

@router.get("/export/csv/")
async def export_csv(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    material_name: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    """CSV出力（UTF-8 BOM付き）"""
    # 集計データ取得（summary APIを流用）
    summary = await get_analytics_summary(
        start_date=start_date,
        end_date=end_date,
        material_name=material_name,
        db=db
    )

    # CSV生成
    output = io.StringIO()
    output.write('\ufeff')  # BOM
    writer = csv.writer(output)

    # ヘッダー
    writer.writerow([
        '材料名', '現在在庫本数', '現在在庫重量(kg)',
        '入庫本数', '入庫重量(kg)', '出庫本数', '出庫重量(kg)', '金額（円）'
    ])

    # データ行
    for m in summary.materials:
        writer.writerow([
            m.material_name,
            m.current_stock_quantity,
            m.current_stock_weight_kg,
            m.total_in_quantity,
            m.total_in_weight_kg,
            m.total_out_quantity,
            m.total_out_weight_kg,
            m.total_amount
        ])

    # 合計行
    writer.writerow([])
    writer.writerow([
        '合計',
        summary.total_stock_quantity,
        summary.total_stock_weight_kg,
        summary.total_in_quantity,
        summary.total_in_weight_kg,
        summary.total_out_quantity,
        summary.total_out_weight_kg,
        summary.total_amount
    ])

    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f"attachment; filename=analytics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        }
    )

@router.get("/export/excel/")
async def export_excel(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    material_name: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    """Excel出力（openpyxl使用）"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxlがインストールされていません")

    # 集計データ取得
    summary = await get_analytics_summary(
        start_date=start_date,
        end_date=end_date,
        material_name=material_name,
        db=db
    )

    # Excel生成
    wb = Workbook()
    ws = wb.active
    ws.title = "集計結果"

    # ヘッダー
    headers = [
        '材料名', '現在在庫本数', '現在在庫重量(kg)',
        '入庫本数', '入庫重量(kg)', '出庫本数', '出庫重量(kg)', '金額（円）'
    ]
    ws.append(headers)

    # ヘッダースタイル
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # データ行
    for m in summary.materials:
        ws.append([
            m.material_name,
            m.current_stock_quantity,
            m.current_stock_weight_kg,
            m.total_in_quantity,
            m.total_in_weight_kg,
            m.total_out_quantity,
            m.total_out_weight_kg,
            m.total_amount
        ])

    # 合計行
    ws.append([])
    total_row = [
        '合計',
        summary.total_stock_quantity,
        summary.total_stock_weight_kg,
        summary.total_in_quantity,
        summary.total_in_weight_kg,
        summary.total_out_quantity,
        summary.total_out_weight_kg,
        summary.total_amount
    ]
    ws.append(total_row)

    # 合計行スタイル
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    # 列幅調整
    ws.column_dimensions['A'].width = 30
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 15

    # バイナリ出力
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=analytics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        }
    )
